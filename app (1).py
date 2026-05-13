import streamlit as st
from playwright.sync_api import sync_playwright
from openpyxl import load_workbook
from datetime import datetime
import os, sys, threading, time
import PyPDF2

# ─── AYARLAR ──────────────────────────────────────────────────────────────────
if getattr(sys, 'frozen', False):
    base = os.path.dirname(sys.executable)
else:
    base = os.path.dirname(os.path.abspath(__file__))

excel_yolu      = os.path.join(base, "veriler.xlsx")
akct_dosya_yolu = os.path.join(base, "akct.xlsx")
pdf_klasor      = base
os.makedirs(pdf_klasor, exist_ok=True)

# ─── SAYFA AYARI ──────────────────────────────────────────────────────────────
st.set_page_config(
    page_title="TARIC Kontrol Paneli",
    page_icon="🛃",
    layout="wide",
    initial_sidebar_state="collapsed",
)

# ─── CSS ──────────────────────────────────────────────────────────────────────
st.markdown("""
<style>
@import url('https://fonts.googleapis.com/css2?family=JetBrains+Mono:wght@400;700&family=Syne:wght@400;700;800&display=swap');

html, body, [data-testid="stAppViewContainer"] {
    background: #0b0e14;
    font-family: 'Syne', sans-serif;
    color: #e2e8f0;
}
[data-testid="stAppViewContainer"] > .main { padding: 0 !important; }
section[data-testid="stSidebar"] { display: none; }

.stButton > button {
    width: 100% !important;
    font-family: 'Syne', sans-serif !important;
    font-weight: 700 !important;
    font-size: 14px !important;
    border-radius: 10px !important;
    padding: 14px 0 !important;
    border: none !important;
    cursor: pointer !important;
    transition: filter 0.15s, transform 0.1s !important;
    letter-spacing: 0.5px !important;
    color: white !important;
}
.stButton > button:hover { filter: brightness(1.15) !important; transform: translateY(-1px) !important; }
.stButton > button:active { transform: translateY(0px) !important; }

.card {
    background: #161b27;
    border: 1px solid #1e2535;
    border-radius: 12px;
    padding: 16px;
    margin-bottom: 10px;
}
.gtip-big {
    font-family: 'JetBrains Mono', monospace;
    font-size: 20px;
    font-weight: 700;
    color: #f8fafc;
    letter-spacing: 2px;
}
.badge-akct  { background:#065f46; color:#6ee7b7; border:1px solid #059669; }
.badge-yoksa { background:#7f1d1d; color:#fca5a5; border:1px solid #dc2626; }
.badge {
    display:inline-block; padding:5px 14px; border-radius:20px;
    font-size:12px; font-weight:700; margin-top:8px;
}
.veri-ok       { background:#14532d; color:#86efac; border:1px solid #22c55e; border-radius:8px; padding:10px; text-align:center; font-weight:700; margin:6px 0; }
.veri-gecersiz { background:#7f1d1d; color:#fca5a5; border:1px solid #ef4444; border-radius:8px; padding:10px; text-align:center; font-weight:700; margin:6px 0; }
.veri-yok      { background:#78350f; color:#fcd34d; border:1px solid #f59e0b; border-radius:8px; padding:10px; text-align:center; font-weight:700; margin:6px 0; }
.veri-bekle    { background:#1e2a3a; color:#94a3b8; border:1px solid #334155; border-radius:8px; padding:10px; text-align:center; font-weight:700; margin:6px 0; }

.log-box {
    background:#0a0c10; border:1px solid #1e2535; border-radius:8px;
    padding:10px; font-family:'JetBrains Mono',monospace; font-size:11px;
    max-height:150px; overflow-y:auto;
}
.log-row { padding:2px 0; border-bottom:1px solid #161b27; color:#64748b; }
.log-row .ts  { color:#38bdf8; }
.log-row .act { color:#4ade80; }
.log-row .cod { color:#fb923c; }

.prog-bar-bg { background:#1e2535; border-radius:8px; height:8px; overflow:hidden; margin:4px 0 10px; }
.prog-bar-fg { background:linear-gradient(90deg,#38bdf8,#818cf8); height:100%; border-radius:8px; transition:width 0.4s; }
hr { border-color:#1e2535 !important; margin:8px 0 !important; }
</style>
""", unsafe_allow_html=True)

# ─── SESSION STATE ─────────────────────────────────────────────────────────────
def init_state():
    defaults = {
        "baslatildi":   False,
        "mevcut_satir": 2,
        "toplam_satir": 0,
        "islenen":      0,
        "gtip_noktalı": "",
        "gtip_12":      "",
        "ulke_ad":      "",
        "ulke_sembol":  "",
        "tarih_str":    "",
        "akct_durumu":  "",
        "veri_mesaji":  "",
        "karar":        None,
        "bekliyor":     False,
        "bitti":        False,
        "current_code": "",
        "pdf_sayisi":   0,
        "hata":         "",
        "log":          [],
    }
    for k, v in defaults.items():
        if k not in st.session_state:
            st.session_state[k] = v

init_state()

# ─── YARDIMCI FONKSİYONLAR ────────────────────────────────────────────────────
def log_ekle(aksiyon, kod=""):
    ts = datetime.now().strftime("%H:%M:%S")
    st.session_state.log.insert(0, {"ts": ts, "act": aksiyon, "cod": str(kod)})
    if len(st.session_state.log) > 50:
        st.session_state.log = st.session_state.log[:50]

def akct_kontrol(gtip_noktalı):
    """GTİP noktaları temizlenmiş 12 haneli kod ile AKÇT listesini karşılaştırır."""
    gtip_temiz = str(gtip_noktalı).replace(".", "").strip()
    if not os.path.exists(akct_dosya_yolu):
        return "⚠️ akct.xlsx Yok"
    try:
        wb = load_workbook(akct_dosya_yolu, data_only=True, read_only=True)
        ws = wb.active
        for row in ws.iter_rows(min_col=1, max_col=1, values_only=True):
            val = str(row[0] or "").replace(".", "").strip()
            if val == gtip_temiz:
                return "🛡️ AKÇT Ürünüdür"
        return "📄 AKÇT değildir"
    except Exception as e:
        return f"Hata: {e}"

def veri_mesaji_belirle(icerik):
    low = icerik.lower()
    if "not or no longer valid" in low:
        return "gecersiz"
    elif "No data matches" in icerik:
        return "yok"
    return "ok"

def tarih_formatla(tarih_val):
    """Excel'den gelen datetime'ı DD-MM-YYYY formatına çevirir."""
    if isinstance(tarih_val, datetime):
        return tarih_val.strftime("%d-%m-%Y")
    s = str(tarih_val).split()[0]
    return s.replace(".", "-").replace("/", "-")

def pdf_birlestir():
    cikti_yolu = os.path.join(pdf_klasor, "Birleştirilmiş.pdf")
    if os.path.exists(cikti_yolu):
        try:
            os.remove(cikti_yolu)
        except:
            return "❌ Birleştirilmiş.pdf açık, önce kapatın!"
    try:
        merger = PyPDF2.PdfMerger()
        pdf_listesi = sorted(
            [f for f in os.listdir(pdf_klasor) if f.endswith(".pdf") and f.startswith("verisirasi_")],
            key=lambda x: int(x.split('_')[1])
        )
        if not pdf_listesi:
            return "⚠️ Birleştirilecek PDF bulunamadı!"
        for pdf in pdf_listesi:
            merger.append(os.path.join(pdf_klasor, pdf))
        merger.write(cikti_yolu)
        merger.close()
        return f"✅ {len(pdf_listesi)} PDF birleştirildi → Birleştirilmiş.pdf"
    except Exception as e:
        return f"❌ Hata: {e}"

# ─── PLAYWRIGHT THREAD ────────────────────────────────────────────────────────
def playwright_thread():
    """
    veriler.xlsx sütun yapısı:
      A (col 1) = GTİP Kodu (noktalı, 12 li)   → AKÇT kontrolü için
      B (col 2) = Menşe Ülke Adı (Türkçe)
      C (col 3) = Tarih (datetime)
      D (col 4) = Goods code  (formül → TARIC sorgu kodu, örn: 7207121000)
      E (col 5) = Origin/dest (formül → ülke sembolü, örn: RU)
      F-H       = Ülke çeviri tablosu (sabit, işlenmez)
    Sadece A+B+C dolu VE D geçerli olan satırlar işlenir.
    """
    try:
        wb = load_workbook(excel_yolu, data_only=True)
        ws = wb.active

        satirlar = []
        for r in range(2, ws.max_row + 1):
            gtip_noktalı = ws.cell(row=r, column=1).value
            ulke_ad      = ws.cell(row=r, column=2).value
            tarih_val    = ws.cell(row=r, column=3).value
            sorgu_kodu   = ws.cell(row=r, column=4).value
            ulke_sembol  = ws.cell(row=r, column=5).value

            # A, B, C dolu olmalı
            if not gtip_noktalı or not ulke_ad or not tarih_val:
                continue
            # D geçerli sorgu kodu olmalı
            sorgu = str(sorgu_kodu or "").strip()
            if not sorgu or sorgu in ("None", "#N/A", ""):
                continue

            satirlar.append({
                "row":         r,
                "gtip_noktalı": str(gtip_noktalı),
                "ulke_ad":     str(ulke_ad),
                "tarih_val":   tarih_val,
                "sorgu":       sorgu,
                "ulke_sembol": str(ulke_sembol or "").strip(),
            })

        st.session_state.toplam_satir = len(satirlar)

        with sync_playwright() as p:
            browser = p.chromium.launch(
                headless=False,
                args=["--start-maximized", "--window-position=380,0"]
            )
            context = browser.new_context(no_viewport=True)
            page    = context.new_page()
            client  = page.context.new_cdp_session(page)
            zoom    = 0.75

            for idx, s in enumerate(satirlar):
                # Session state güncelle
                st.session_state.mevcut_satir  = s["row"]
                st.session_state.islenen       = idx
                st.session_state.gtip_noktalı  = s["gtip_noktalı"]
                st.session_state.gtip_12       = s["gtip_noktalı"].replace(".", "")
                st.session_state.ulke_ad       = s["ulke_ad"]
                st.session_state.ulke_sembol   = s["ulke_sembol"]
                st.session_state.tarih_str     = tarih_formatla(s["tarih_val"])
                st.session_state.akct_durumu   = akct_kontrol(s["gtip_noktalı"])

                current_code = s["sorgu"]
                st.session_state.current_code  = current_code

                while len(current_code) >= 4:
                    try:
                        page.goto("https://ec.europa.eu/taxation_customs/dds2/taric/taric_consultation.jsp?Lang=en")
                        client.send("Emulation.setPageScaleFactor", {"pageScaleFactor": zoom})

                        page.fill("#taricCode", current_code)

                        ulke = s["ulke_sembol"]
                        if ulke and ulke not in ("None", "#N/A", ""):
                            try:
                                page.select_option("#taricArea", ulke)
                            except:
                                pass

                        page.evaluate(f"document.querySelector('#SimDatePic').value = '{st.session_state.tarih_str}'")
                        page.click("button[value='Retrieve Measures']")
                        page.wait_for_load_state("networkidle")
                        client.send("Emulation.setPageScaleFactor", {"pageScaleFactor": zoom})

                        icerik = page.content()
                        st.session_state.veri_mesaji  = veri_mesaji_belirle(icerik)
                        st.session_state.current_code = current_code
                        st.session_state.karar        = None
                        st.session_state.bekliyor     = True

                        # Kullanıcı kararını bekle
                        while st.session_state.bekliyor:
                            time.sleep(0.3)

                        karar = st.session_state.karar

                        if karar == "kirp":
                            current_code = current_code[:-2]
                            st.session_state.current_code = current_code
                            log_ekle("Kod Kısaltıldı", current_code)

                        elif karar == "pdf":
                            dosya = os.path.join(pdf_klasor, f"verisirasi_{idx+1}_{current_code}.pdf")
                            page.pdf(path=dosya, format="A4", print_background=True)
                            st.session_state.pdf_sayisi += 1
                            log_ekle("PDF Alındı", current_code)
                            break

                        elif karar == "atla":
                            log_ekle("Atlandı", current_code)
                            break

                        elif karar == "bitir":
                            browser.close()
                            st.session_state.bitti    = True
                            st.session_state.bekliyor = False
                            return

                    except Exception as e:
                        st.session_state.hata    = str(e)
                        st.session_state.bekliyor = False
                        log_ekle(f"HATA", str(e)[:60])
                        break

            browser.close()
            st.session_state.bitti    = True
            st.session_state.islenen  = len(satirlar)
            st.session_state.bekliyor = False
            log_ekle("✅ Tüm Satırlar Tamamlandı")

    except PermissionError:
        st.session_state.hata     = "veriler.xlsx açık! Lütfen Excel'i kapatın."
        st.session_state.bekliyor = False
    except Exception as e:
        st.session_state.hata     = str(e)
        st.session_state.bekliyor = False

# ─── BUTON CALLBACK'LERİ ──────────────────────────────────────────────────────
def karar_ver(k):
    st.session_state.karar    = k
    st.session_state.bekliyor = False

# ─── ARAYÜZ ───────────────────────────────────────────────────────────────────
sol, sag = st.columns([1, 2], gap="medium")

with sol:
    st.markdown("""
    <div style='padding:20px 16px 12px;border-bottom:1px solid #1e2535;margin-bottom:12px;'>
        <div style='font-family:Syne,sans-serif;font-size:24px;font-weight:800;
                    background:linear-gradient(90deg,#38bdf8,#818cf8);
                    -webkit-background-clip:text;-webkit-text-fill-color:transparent;'>
            🛃 TARIC Panel
        </div>
        <div style='font-size:11px;color:#4b5563;'>İnsan Kontrollü Otomasyon</div>
    </div>
    """, unsafe_allow_html=True)

    # ── BAŞLATILMADI ──────────────────────────────────────────────────────────
    if not st.session_state.baslatildi:
        if not os.path.exists(excel_yolu):
            st.error(f"❌ veriler.xlsx bulunamadı!\n`{excel_yolu}`")
        else:
            st.markdown("""
            <div class='card' style='text-align:center;padding:28px 16px;'>
                <div style='font-size:44px;'>🛃</div>
                <div style='font-size:13px;color:#94a3b8;margin:12px 0 4px;line-height:1.7;'>
                    veriler.xlsx hazır.<br>
                    Başlat'a basınca Chromium açılır<br>
                    ve TARIC sorgusunu yapar.
                </div>
            </div>
            """, unsafe_allow_html=True)
            if st.button("🚀  Sorgulamayı Başlat", use_container_width=True):
                st.session_state.baslatildi = True
                threading.Thread(target=playwright_thread, daemon=True).start()
                st.rerun()

    # ── ÇALIŞIYOR ─────────────────────────────────────────────────────────────
    elif not st.session_state.bitti:

        islenen = st.session_state.islenen
        toplam  = max(st.session_state.toplam_satir, 1)
        pct     = int((islenen / toplam) * 100)

        st.markdown(f"""
        <div style='font-size:11px;color:#4b5563;font-family:JetBrains Mono,monospace;'>
            {islenen} / {toplam} satır &nbsp;·&nbsp; %{pct} &nbsp;·&nbsp; 📄 {st.session_state.pdf_sayisi} PDF
        </div>
        <div class='prog-bar-bg'><div class='prog-bar-fg' style='width:{pct}%'></div></div>
        """, unsafe_allow_html=True)

        # GTİP Kartı
        gtip    = st.session_state.gtip_noktalı or "—"
        kod     = st.session_state.current_code  or "—"
        ulke_ad = st.session_state.ulke_ad       or "—"
        sembol  = st.session_state.ulke_sembol   or "—"
        tarih   = st.session_state.tarih_str     or "—"
        akct    = st.session_state.akct_durumu   or "—"
        akct_cls = "badge-akct" if "Ürünüdür" in akct else "badge-yoksa"

        st.markdown(f"""
        <div class='card'>
            <div style='font-size:10px;color:#4b5563;margin-bottom:4px;letter-spacing:1px;'>AKTİF GTİP</div>
            <div class='gtip-big'>{gtip}</div>
            <div style='margin-top:10px;font-size:12px;color:#64748b;
                        line-height:2;font-family:JetBrains Mono,monospace;'>
                🔢 Sorgu : <b style='color:#e2e8f0;'>{kod}</b><br>
                🌍 Ülke  : <b style='color:#e2e8f0;'>{ulke_ad}</b> ({sembol})<br>
                📅 Tarih : <b style='color:#e2e8f0;'>{tarih}</b>
            </div>
            <div><span class='badge {akct_cls}'>{akct}</span></div>
        </div>
        """, unsafe_allow_html=True)

        # Veri Durumu
        vm = st.session_state.veri_mesaji
        if   vm == "ok":       st.markdown("<div class='veri-ok'>✅ Veri OK — Sayfayı inceleyin</div>", unsafe_allow_html=True)
        elif vm == "gecersiz": st.markdown("<div class='veri-gecersiz'>❌ GEÇERSİZ KOD</div>", unsafe_allow_html=True)
        elif vm == "yok":      st.markdown("<div class='veri-yok'>⚠️ VERİ YOK</div>", unsafe_allow_html=True)
        else:                  st.markdown("<div class='veri-bekle'>⏳ Yükleniyor...</div>", unsafe_allow_html=True)

        st.markdown("<div style='height:4px'></div>", unsafe_allow_html=True)

        # Ana butonlar
        if st.session_state.bekliyor:
            c1, c2 = st.columns(2)
            with c1:
                if st.button("🖨️ ONAYLA\n(PDF AL)", use_container_width=True):
                    karar_ver("pdf"); st.rerun()
            with c2:
                if st.button("⏭️ SIRADAKİNE\nGEÇ", use_container_width=True):
                    karar_ver("atla"); st.rerun()

            st.markdown("<div style='height:4px'></div>", unsafe_allow_html=True)
            if st.button("✂️  KODU AZALT  (−2 karakter)", use_container_width=True):
                karar_ver("kirp"); st.rerun()
        else:
            st.markdown("<div class='veri-bekle'>⏳ Sayfa yükleniyor...</div>", unsafe_allow_html=True)
            time.sleep(0.5)
            st.rerun()

        st.markdown("<hr>", unsafe_allow_html=True)

        if st.button("📎  PDF BİRLEŞTİR", use_container_width=True):
            st.success(pdf_birlestir())
            log_ekle("PDF Birleştirme")

        if st.button("🔴  SİSTEMİ KAPAT", use_container_width=True):
            karar_ver("bitir"); st.rerun()

        st.markdown("<hr>", unsafe_allow_html=True)

        # Log
        st.markdown("<div style='font-size:11px;color:#4b5563;margin-bottom:4px;'>📋 İşlem Geçmişi</div>", unsafe_allow_html=True)
        if st.session_state.log:
            rows = "".join(
                f"<div class='log-row'><span class='ts'>[{e['ts']}]</span>"
                f"<span class='act'> {e['act']}</span>"
                f"<span class='cod'> {e['cod']}</span></div>"
                for e in st.session_state.log
            )
            st.markdown(f"<div class='log-box'>{rows}</div>", unsafe_allow_html=True)

        if st.session_state.hata:
            st.error(f"⚠️ {st.session_state.hata}")

        if not st.session_state.bekliyor and not st.session_state.bitti:
            time.sleep(0.4)
            st.rerun()

    # ── BİTTİ ─────────────────────────────────────────────────────────────────
    else:
        st.success(f"🎉 Tüm {st.session_state.toplam_satir} satır tamamlandı!")
        st.markdown(f"**Toplam PDF:** {st.session_state.pdf_sayisi}")

        if st.button("📎  PDF BİRLEŞTİR", use_container_width=True):
            st.success(pdf_birlestir())

        if st.button("🔄  Yeniden Başlat", use_container_width=True):
            for k in list(st.session_state.keys()):
                del st.session_state[k]
            st.rerun()

# ─── SAĞ PANEL ────────────────────────────────────────────────────────────────
with sag:
    if st.session_state.baslatildi and not st.session_state.bitti:
        bekliyor   = st.session_state.bekliyor
        durum_renk = "#065f46" if bekliyor else "#1a2236"
        durum_yazi = "🟢 Sayfayı inceleyin → Sol panelden kararınızı verin" if bekliyor else "⏳ TARIC sayfası yükleniyor..."

        st.markdown(f"""
        <div style='margin:24px 24px 16px;padding:18px 22px;background:{durum_renk};
                    border-radius:12px;border:1px solid #1e2535;font-weight:700;font-size:15px;'>
            {durum_yazi}
        </div>
        """, unsafe_allow_html=True)

        st.markdown("""
        <div style='margin:0 24px;padding:22px;background:#111520;border:1px solid #1e2535;border-radius:12px;'>
            <div style='font-size:12px;font-weight:700;color:#475569;margin-bottom:14px;letter-spacing:1px;'>KULLANIM AKIŞI</div>
            <div style='font-size:13px;color:#64748b;line-height:2.2;'>
                1️⃣ &nbsp;Chromium penceresinde TARIC sayfasını inceleyin<br>
                2️⃣ &nbsp;<b style='color:#94a3b8;'>▶ ok işaretlerine tıklayarak</b> detayları açın<br>
                3️⃣ &nbsp;Sayfa hazırsa &nbsp;→&nbsp; <b style='color:#4ade80;'>🖨️ ONAYLA (PDF AL)</b><br>
                4️⃣ &nbsp;Kod hatalıysa &nbsp;→&nbsp; <b style='color:#fb923c;'>✂️ KODU AZALT</b> (−2 karakter)<br>
                5️⃣ &nbsp;Bu satırı geç &nbsp;→&nbsp; <b style='color:#38bdf8;'>⏭️ SIRADAKİNE GEÇ</b><br>
                6️⃣ &nbsp;Son adım &nbsp;→&nbsp; <b style='color:#c084fc;'>📎 PDF BİRLEŞTİR</b>
            </div>
        </div>
        """, unsafe_allow_html=True)

        st.markdown("""
        <div style='margin:14px 24px 0;padding:18px 20px;background:#111520;border:1px solid #1e2535;border-radius:12px;'>
            <div style='font-size:12px;font-weight:700;color:#475569;margin-bottom:10px;letter-spacing:1px;'>VERİLER.XLSX SÜTUN YAPISI</div>
            <div style='font-size:12px;color:#64748b;font-family:JetBrains Mono,monospace;line-height:2;'>
                A → GTİP Kodu (noktalı 12 li)  — AKÇT kontrolü<br>
                B → Menşe Ülke Adı (Türkçe)<br>
                C → Tarih (referans tarihi)<br>
                D → Goods code (formül → TARIC sorgu kodu)<br>
                E → Ülke sembolü (formül → RU, DE, CN...)<br>
                F–H → Ülke çeviri tablosu (sabit, kodda kullanılmaz)
            </div>
        </div>
        """, unsafe_allow_html=True)

    elif not st.session_state.baslatildi:
        st.markdown("""
        <div style='margin:40px 24px;padding:52px;background:#111520;border:2px dashed #1e2535;
                    border-radius:16px;text-align:center;'>
            <div style='font-size:56px;margin-bottom:16px;'>🛃</div>
            <div style='font-size:20px;font-weight:800;color:#94a3b8;margin-bottom:10px;'>
                TARIC Kontrol Paneli
            </div>
            <div style='font-size:13px;color:#4b5563;line-height:2;'>
                Sol panelden <b style='color:#38bdf8;'>Sorgulamayı Başlat</b>'a basın.<br>
                Chromium otomatik açılır, TARIC sorgusunu yapar.<br>
                Siz sayfayı inceleyin, sol panelden kararınızı verin.
            </div>
        </div>
        """, unsafe_allow_html=True)
